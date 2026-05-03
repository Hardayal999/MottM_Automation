from copy import copy
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter


def parse_columns(text):
    """Parse A,D,G,J or 1,4,7,10 into sorted 1-based column numbers."""
    if not text:
        return None
    cols = []
    for part in text.split(","):
        token = part.strip()
        if not token:
            continue
        cols.append(column_index_from_string(token.upper()) if token.isalpha() else int(token))
    return sorted(set(cols))


def column_letters(cols):
    return ",".join(get_column_letter(c) for c in cols)


def normalise_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def is_empty_row(values):
    return all(v is None or str(v).strip() == "" for v in values)


def copy_cell(src, dst):
    """Copy value and style from one openpyxl cell to another."""
    dst.value = src.value
    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.protection = copy(src.protection)
    dst.font = copy(src.font)


def copy_sheet_layout(src_ws, dst_ws, max_col):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        dst_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines


def copy_row_style(src_ws, dst_ws, src_row, dst_row, max_col):
    dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        copy_cell(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col))


def clone_font_with(cell, *, color=None, strike=None):
    base = copy(cell.font) if cell.font else Font()
    if color is not None:
        base.color = color
    if strike is not None:
        base.strike = strike
    return base


def apply_change_style(cell, fill_hex, font_hex, strike=False):
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    cell.font = clone_font_with(cell, color=font_hex, strike=strike)


def soften_repeated_root_cell(cell, font_hex="A6A6A6", border_hex="D9D9D9"):
    """
    Visually fade repeated top/root IDs while keeping the right border unchanged,
    so the cell still connects neatly to the next column.
    """
    cell.font = clone_font_with(cell, color=font_hex)
    thin = Side(style="thin", color=border_hex)
    existing = cell.border or Border()
    cell.border = Border(
        left=thin,
        top=thin,
        bottom=thin,
        right=existing.right,
        diagonal=existing.diagonal,
        diagonal_direction=existing.diagonal_direction,
        diagonalUp=existing.diagonalUp,
        diagonalDown=existing.diagonalDown,
        outline=existing.outline,
        vertical=existing.vertical,
        horizontal=existing.horizontal,
    )
