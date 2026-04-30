from openpyxl.utils import get_column_letter
from .excel_utils import copy_row_style, copy_sheet_layout, apply_change_style, soften_repeated_root_cell


def write_output(wb, new_ws, old_ws, output_records, max_col, config, id_cols=None):
    if config.output_sheet_name in wb.sheetnames:
        del wb[config.output_sheet_name]
    ws = wb.create_sheet(config.output_sheet_name, 0)
    copy_sheet_layout(new_ws, ws, max_col)
    for row in range(1, config.header_rows + 1):
        copy_row_style(new_ws, ws, row, row, max_col)
    out_row = config.header_rows + 1
    for item in output_records:
        source_ws = old_ws if item.status == "deleted" else new_ws
        copy_row_style(source_ws, ws, item.record.source_row, out_row, max_col)
        for col, value in enumerate(item.record.values, start=1):
            ws.cell(out_row, col).value = value
        if item.status == "added":
            for col in range(item.change_start_col, max_col + 1):
                apply_change_style(ws.cell(out_row, col), config.added_fill, config.added_font, False)
        elif item.status == "deleted":
            for col in range(item.change_start_col, max_col + 1):
                apply_change_style(ws.cell(out_row, col), config.deleted_fill, config.deleted_font, True)
        out_row += 1
    if id_cols:
        apply_repeated_root_id_format(
            ws,
            config.header_rows + 1,
            out_row - 1,
            id_cols[0],
            config.repeated_root_font,
            config.repeated_root_border,
        )
    ws.auto_filter.ref = f"A{config.header_rows}:{get_column_letter(max_col)}{max(out_row - 1, config.header_rows)}"
    return ws


def apply_repeated_root_id_format(ws, first_data_row, last_data_row, root_col, font_hex, border_hex):
    previous = None
    for row in range(first_data_row, last_data_row + 1):
        cell = ws.cell(row, root_col)
        current = str(cell.value).strip() if cell.value is not None else ""
        if current and current == previous:
            soften_repeated_root_cell(cell, font_hex, border_hex)
        elif current:
            previous = current
        else:
            previous = current
