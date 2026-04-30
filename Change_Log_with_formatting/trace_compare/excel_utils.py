from copy import copy
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font


def parse_columns(text):
    if not text:
        return None
    cols = []
    for part in text.split(","):
        token = part.strip()
        if not token:
            continue
        cols.append(column_index_from_string(token.upper()) if token.isalpha() else int(token))
    return sorted(set(cols))


def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.border:
        dst.border = copy(src.border)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.protection:
        dst.protection = copy(src.protection)
    if src.font:
        dst.font = copy(src.font)


def copy_row_style(src_ws, dst_ws, src_row, dst_row, max_col):
    dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        copy_cell(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col))


def copy_sheet_layout(src_ws, dst_ws, max_col):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        dst_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines


def apply_change_style(cell, fill_hex, font_hex, strike=False):
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    base = cell.font.copy() if cell.font else Font()
    cell.font = Font(
        name=base.name,
        sz=base.sz,
        b=base.b,
        i=base.i,
        vertAlign=base.vertAlign,
        underline=base.underline,
        strike=strike,
        color=font_hex,
        charset=base.charset,
        family=base.family,
        scheme=base.scheme,
        outline=base.outline,
        shadow=base.shadow,
        condense=base.condense,
        extend=base.extend,
    )


def is_empty_row(values):
    return all(v is None or str(v).strip() == "" for v in values)

from openpyxl.styles import Side, Border

def soften_repeated_root_cell(cell, font_hex="A6A6A6", border_hex="D9D9D9"):
    base = cell.font.copy() if cell.font else Font()
    cell.font = Font(
        name=base.name,
        sz=base.sz,
        b=base.b,
        i=base.i,
        vertAlign=base.vertAlign,
        underline=base.underline,
        strike=base.strike,
        color=font_hex,
        charset=base.charset,
        family=base.family,
        scheme=base.scheme,
        outline=base.outline,
        shadow=base.shadow,
        condense=base.condense,
        extend=base.extend,
    )
    light = Side(style="thin", color=border_hex)
    existing = cell.border.copy() if cell.border else Border()
    cell.border = Border(
        left=light,
        right=existing.right,
        top=light,
        bottom=light,
        diagonal=existing.diagonal,
        diagonal_direction=existing.diagonal_direction,
        diagonalUp=existing.diagonalUp,
        diagonalDown=existing.diagonalDown,
        outline=existing.outline,
        vertical=existing.vertical,
        horizontal=existing.horizontal,
    )
