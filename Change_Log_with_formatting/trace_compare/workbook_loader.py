from openpyxl import load_workbook


def load_pair(path, new_suffix, old_suffix):
    wb = load_workbook(path)
    new_sheets = [s for s in wb.sheetnames if s.lower().endswith(new_suffix.lower())]
    old_sheets = [s for s in wb.sheetnames if s.lower().endswith(old_suffix.lower())]
    if len(new_sheets) != 1 or len(old_sheets) != 1:
        raise ValueError(f"Expected exactly one sheet ending {new_suffix!r} and one ending {old_suffix!r}. Found new={new_sheets}, old={old_sheets}")
    return wb, wb[new_sheets[0]], wb[old_sheets[0]]
