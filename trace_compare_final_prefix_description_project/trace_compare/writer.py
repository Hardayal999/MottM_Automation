from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import CompareConfig
from .diff_engine import added_start_col, deleted_start_col
from .excel_utils import apply_fill, apply_repeated_root_style, copy_cell_style, normalise_value
from .row_model import RowRecord


def remove_existing_output_sheet(wb, output_sheet_name: str) -> None:
    if output_sheet_name in wb.sheetnames:
        del wb[output_sheet_name]


def copy_sheet_layout(src_ws: Worksheet, dst_ws: Worksheet, max_col: int) -> None:
    """Copy column widths, row heights, freeze panes and page settings."""
    for col_idx in range(1, max_col + 1):
        letter = src_ws.cell(1, col_idx).column_letter
        if letter in src_ws.column_dimensions:
            dst_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width
            dst_ws.column_dimensions[letter].hidden = src_ws.column_dimensions[letter].hidden

    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height
        dst_ws.row_dimensions[row_idx].hidden = dim.hidden

    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines


def write_header(src_ws: Worksheet, out_ws: Worksheet, max_col: int, config: CompareConfig) -> None:
    for col in range(1, max_col + 1):
        src = src_ws.cell(config.header_row, col)
        dst = out_ws.cell(config.header_row, col)
        dst.value = src.value
        copy_cell_style(src, dst)


def write_record_row(
    out_ws: Worksheet,
    new_ws: Worksheet,
    old_ws: Worksheet,
    record: RowRecord,
    out_row: int,
    max_col: int,
) -> None:
    source_ws = new_ws if record.status != "deleted" else old_ws
    source_row = record.source_row

    for col in range(1, max_col + 1):
        src = source_ws.cell(source_row, col)
        dst = out_ws.cell(out_row, col)
        # If a hierarchy prefix description/detail changed, display the newer
        # value in the output and highlight it yellow later. This matters even
        # on deleted relationship rows where the parent still exists and only
        # the child link was deleted.
        if col in record.changed_detail_values:
            dst.value = record.changed_detail_values[col]
        else:
            dst.value = record.values.get(col)
        copy_cell_style(src, dst)

    # Copy original source row height if available.
    source_dim = source_ws.row_dimensions.get(source_row)
    if source_dim and source_dim.height:
        out_ws.row_dimensions[out_row].height = source_dim.height


def apply_status_formatting(
    out_ws: Worksheet,
    records: List[RowRecord],
    old_prefix_sets,
    new_prefix_sets,
    id_columns: List[int],
    max_col: int,
    config: CompareConfig,
) -> None:
    for idx, record in enumerate(records, start=config.data_start_row):
        protected_from_yellow_start = max_col + 1

        if record.status == "added":
            start_col = added_start_col(record, old_prefix_sets, id_columns)
            protected_from_yellow_start = start_col
            apply_fill(out_ws, idx, start_col, max_col, config.added_fill, strike=False)
        elif record.status == "deleted":
            start_col = deleted_start_col(record, new_prefix_sets, id_columns)
            protected_from_yellow_start = start_col
            apply_fill(out_ws, idx, start_col, max_col, config.deleted_fill, strike=True)

        # Prefix-level description/detail changes are yellow only when they are
        # outside the green/red region. This preserves priority:
        # green added > red deleted > yellow changed detail.
        for col_idx in record.changed_detail_columns or []:
            if col_idx < protected_from_yellow_start:
                apply_fill(out_ws, idx, col_idx, col_idx, config.changed_fill, strike=False)


def apply_repeated_root_grey(out_ws: Worksheet, id_columns: List[int], config: CompareConfig) -> None:
    """Grey out consecutive duplicate root IDs for easier scanning."""
    if not id_columns:
        return
    root_col = id_columns[0]
    previous = None

    for row in range(config.data_start_row, out_ws.max_row + 1):
        cell = out_ws.cell(row, root_col)
        current = normalise_value(cell.value)
        if current and current == previous:
            apply_repeated_root_style(cell, config.repeated_root_fill, config.light_border)
        elif current:
            previous = current


def write_output_workbook(
    wb,
    new_ws: Worksheet,
    old_ws: Worksheet,
    records: List[RowRecord],
    old_prefix_sets,
    new_prefix_sets,
    id_columns: List[int],
    max_col: int,
    output_path: str,
    config: CompareConfig,
) -> str:
    remove_existing_output_sheet(wb, config.output_sheet_name)
    out_ws = wb.create_sheet(config.output_sheet_name, 0)

    copy_sheet_layout(new_ws, out_ws, max_col)
    write_header(new_ws, out_ws, max_col, config)

    for out_row, record in enumerate(records, start=config.data_start_row):
        write_record_row(out_ws, new_ws, old_ws, record, out_row, max_col)

    apply_status_formatting(out_ws, records, old_prefix_sets, new_prefix_sets, id_columns, max_col, config)
    apply_repeated_root_grey(out_ws, id_columns, config)

    # Add autofilter over the populated range without changing the visual format.
    if out_ws.max_row >= config.header_row:
        out_ws.auto_filter.ref = out_ws.dimensions

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return str(output)
