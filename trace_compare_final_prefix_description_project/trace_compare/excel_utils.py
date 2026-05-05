from __future__ import annotations

from copy import copy
from typing import Iterable, List

from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter


def normalise_value(value) -> str:
    """Normalise Excel cell values for reliable comparisons."""
    if value is None:
        return ""
    text = str(value).strip()
    return text


def parse_column_list(raw: str) -> List[int]:
    """Parse A,D,G,J or 1,4,7,10 into sorted column indexes."""
    if not raw:
        raise ValueError("--id-columns is required, for example: --id-columns A,D,G,J")

    cols: List[int] = []
    for part in raw.split(','):
        token = part.strip()
        if not token:
            continue
        if token.isdigit():
            idx = int(token)
        else:
            idx = column_index_from_string(token.upper())
        if idx <= 0:
            raise ValueError(f"Invalid column reference: {token}")
        cols.append(idx)

    if not cols:
        raise ValueError("No valid ID columns were supplied.")

    unique_sorted = sorted(set(cols))
    if unique_sorted != cols:
        raise ValueError("ID columns must be unique and listed left-to-right, e.g. A,D,G,J")
    return cols


def col_letters(cols: Iterable[int]) -> str:
    return ",".join(get_column_letter(c) for c in cols)


def copy_cell_style(src, dst) -> None:
    """Copy visual style from one openpyxl cell to another."""
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def apply_fill(ws, row: int, start_col: int, end_col: int, argb: str, strike: bool = False) -> None:
    fill = PatternFill(fill_type="solid", fgColor=argb)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        if strike:
            f = copy(cell.font)
            f.strike = True
            cell.font = f


def apply_repeated_root_style(cell, fill_argb: str, border_argb: str) -> None:
    """Grey repeated root ID cells and lighten three edges, leaving the right edge untouched."""
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_argb)
    light = Side(style="thin", color=border_argb)
    existing = cell.border or Border()
    cell.border = Border(
        left=light,
        top=light,
        bottom=light,
        right=existing.right,
        diagonal=existing.diagonal,
        diagonal_direction=existing.diagonal_direction,
        diagonalUp=existing.diagonalUp,
        diagonalDown=existing.diagonalDown,
        outline=existing.outline,
        vertical=existing.vertical,
        horizontal=existing.horizontal,
    )


def row_is_blank(values: dict[int, object], max_col: int) -> bool:
    return all(normalise_value(values.get(c)) == "" for c in range(1, max_col + 1))
