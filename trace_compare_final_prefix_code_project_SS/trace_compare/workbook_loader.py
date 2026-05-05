from __future__ import annotations

from pathlib import Path
from typing import Tuple

from openpyxl import load_workbook

from .config import CompareConfig


def load_workbook_and_sheets(input_path: str, config: CompareConfig):
    path = Path(input_path)
    if not path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    wb = load_workbook(path)

    new_sheets = [s for s in wb.sheetnames if s.lower().endswith(config.new_sheet_suffix.lower())]
    old_sheets = [s for s in wb.sheetnames if s.lower().endswith(config.old_sheet_suffix.lower())]

    if len(new_sheets) != 1:
        raise ValueError(f"Expected exactly one sheet ending with {config.new_sheet_suffix}; found {new_sheets}")
    if len(old_sheets) != 1:
        raise ValueError(f"Expected exactly one sheet ending with {config.old_sheet_suffix}; found {old_sheets}")

    return wb, wb[new_sheets[0]], wb[old_sheets[0]]
