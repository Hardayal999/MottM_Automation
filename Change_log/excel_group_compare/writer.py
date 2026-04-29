from copy import copy
from openpyxl.utils import get_column_letter
from .config import CompareConfig
from .diff_engine import DiffResult
from .excel_utils import copy_cell_style, copy_sheet_layout, apply_added_style, apply_deleted_style
from .preprocess import first_source_row_for_output


def remove_output_sheets(workbook, config: CompareConfig) -> None:
    removable_names = [config.output_sheet_name, "Formatted_Comparison", "Comparison_Summary", "Change_Log", "Deleted_IDs"]
    for name in removable_names:
        if name in workbook.sheetnames:
            del workbook[name]


def _copy_row_from_source(source_ws, target_ws, source_row: int, target_row: int, max_col: int) -> None:
    source_style_row = first_source_row_for_output(source_ws, source_row)
    for col in range(1, max_col + 1):
        source_cell = source_ws.cell(source_style_row, col)
        target_cell = target_ws.cell(target_row, col)
        copy_cell_style(source_cell, target_cell)
    if source_ws.row_dimensions[source_style_row].height:
        target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_style_row].height


def write_comparison_sheet(workbook, new_ws, old_ws, headers: list, diff: DiffResult, config: CompareConfig):
    remove_output_sheets(workbook, config)
    target_ws = workbook.create_sheet(config.output_sheet_name, 0)
    copy_sheet_layout(new_ws, target_ws)
    max_col = len(headers)

    for col in range(1, max_col + 1):
        source_cell = new_ws.cell(config.header_row, col)
        target_cell = target_ws.cell(1, col)
        target_cell.value = headers[col - 1]
        copy_cell_style(source_cell, target_cell)

    target_row = 2
    for diff_row in diff.rows:
        source_ws = new_ws if diff_row.status in ["unchanged", "added"] else old_ws
        _copy_row_from_source(source_ws, target_ws, diff_row.record.source_row, target_row, max_col)
        for col, value in enumerate(diff_row.record.raw_values, start=1):
            cell = target_ws.cell(target_row, col)
            cell.value = value
            if diff_row.status == "added" and value not in (None, ""):
                apply_added_style(cell, config.added_fill, config.added_font)
            elif diff_row.status == "deleted" and value not in (None, ""):
                apply_deleted_style(cell, config.deleted_fill, config.deleted_font)
        target_row += 1

    if max_col > 0 and target_row > 2:
        target_ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{target_row - 1}"
    target_ws.freeze_panes = "A2"
    return target_ws
