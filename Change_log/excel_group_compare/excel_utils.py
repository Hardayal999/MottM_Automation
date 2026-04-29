from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font


_SIMPLE_REF_CHARS = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789$")


def col_to_index(value: str | int) -> int:
    if isinstance(value, int):
        return value
    return column_index_from_string(value.strip().upper())


def parse_column_list(value: str | None) -> list[int] | None:
    if not value:
        return None
    parts = [p.strip() for p in value.split(",") if p.strip()]
    return [col_to_index(p) for p in parts]


def normalise_header(value, case_sensitive: bool = False) -> str:
    text = "" if value is None else str(value).strip()
    return text if case_sensitive else text.lower()


def normalise_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return " ".join(text.split())


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def copy_sheet_layout(source_ws, target_ws) -> None:
    for idx, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[idx].width = dim.width
        target_ws.column_dimensions[idx].hidden = dim.hidden
        target_ws.column_dimensions[idx].outlineLevel = dim.outlineLevel
    for idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx].height = dim.height
        target_ws.row_dimensions[idx].hidden = dim.hidden
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines


def apply_added_style(cell, fill_hex: str, font_hex: str) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    base = copy(cell.font) if cell.font else Font()
    cell.font = Font(
        name=base.name,
        sz=base.sz,
        b=base.b,
        i=base.i,
        u=base.u,
        strike=base.strike,
        color=font_hex,
        vertAlign=base.vertAlign,
        charset=base.charset,
        family=base.family,
        scheme=base.scheme,
        outline=base.outline,
        shadow=base.shadow,
        condense=base.condense,
        extend=base.extend,
    )


def apply_deleted_style(cell, fill_hex: str, font_hex: str) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    base = copy(cell.font) if cell.font else Font()
    cell.font = Font(
        name=base.name,
        sz=base.sz,
        b=base.b,
        i=base.i,
        u=base.u,
        strike=True,
        color=font_hex,
        vertAlign=base.vertAlign,
        charset=base.charset,
        family=base.family,
        scheme=base.scheme,
        outline=base.outline,
        shadow=base.shadow,
        condense=base.condense,
        extend=base.extend,
    )


def is_simple_formula_reference(value) -> bool:
    if not isinstance(value, str) or not value.startswith("="):
        return False
    body = value[1:].replace("$", "").upper()
    return bool(body) and all(ch in _SIMPLE_REF_CHARS for ch in body)


def resolve_simple_reference(ws, value, stack: set[str] | None = None):
    if not is_simple_formula_reference(value):
        return value
    ref = value[1:].replace("$", "")
    stack = stack or set()
    if ref in stack:
        return value
    stack.add(ref)
    referenced = ws[ref].value
    if is_simple_formula_reference(referenced):
        return resolve_simple_reference(ws, referenced, stack)
    return referenced


def cell_coordinate(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"
