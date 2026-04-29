from openpyxl import load_workbook
from .config import CompareConfig


def load_input_workbook(config: CompareConfig):
    return load_workbook(config.input_path)


def find_version_sheets(workbook, config: CompareConfig):
    new_matches = [s for s in workbook.sheetnames if s.lower().endswith(config.new_sheet_suffix.lower())]
    old_matches = [s for s in workbook.sheetnames if s.lower().endswith(config.old_sheet_suffix.lower())]
    if len(new_matches) != 1 or len(old_matches) != 1:
        raise ValueError(
            f"Expected exactly one sheet ending with '{config.new_sheet_suffix}' and one ending with '{config.old_sheet_suffix}'. "
            f"Found new={new_matches}, old={old_matches}."
        )
    return workbook[new_matches[0]], workbook[old_matches[0]]
